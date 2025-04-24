"""
File parsing module for ETL operations.

This module provides utilities for working with various file formats
including CSV, Excel, JSON, XML, and geospatial formats.
"""
import os
import csv
import json
import logging
import tempfile
from typing import Dict, List, Any, Union, Optional, Tuple, Iterator
from datetime import datetime

# Conditional imports based on available modules
try:
    import pandas as pd
    import numpy as np
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    import geopandas as gpd
    HAS_GEOPANDAS = True
except ImportError:
    HAS_GEOPANDAS = False

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xml.etree.ElementTree as ET
    HAS_XML = True
except ImportError:
    HAS_XML = False

# Configure logger
logger = logging.getLogger(__name__)

class FileParser:
    """
    Utility class for parsing various file formats used in ETL processes.
    """
    
    @staticmethod
    def read_csv(file_path: str, 
                 delimiter: str = ',', 
                 has_header: bool = True, 
                 encoding: str = 'utf-8',
                 use_pandas: bool = True) -> Union[List[Dict[str, str]], pd.DataFrame]:
        """
        Read a CSV file into a list of dictionaries or a pandas DataFrame.
        
        Args:
            file_path (str): Path to the CSV file
            delimiter (str): Field delimiter (default: ',')
            has_header (bool): Whether the file has a header row
            encoding (str): File encoding (default: 'utf-8')
            use_pandas (bool): Whether to use pandas if available (default: True)
            
        Returns:
            Union[List[Dict[str, str]], pd.DataFrame]: Parsed data
        """
        if use_pandas and HAS_PANDAS:
            try:
                df = pd.read_csv(file_path, delimiter=delimiter, header=0 if has_header else None, encoding=encoding)
                return df
            except Exception as e:
                logger.warning(f"Error reading CSV with pandas, falling back to csv module: {str(e)}")
                
        # Fallback to standard csv module
        data = []
        try:
            with open(file_path, 'r', encoding=encoding) as csvfile:
                if has_header:
                    reader = csv.DictReader(csvfile, delimiter=delimiter)
                    for row in reader:
                        data.append(dict(row))
                else:
                    reader = csv.reader(csvfile, delimiter=delimiter)
                    for row in reader:
                        data.append(row)
            return data
        except Exception as e:
            logger.exception(f"Error reading CSV file {file_path}: {str(e)}")
            raise
    
    @staticmethod
    def read_excel(file_path: str, 
                  sheet_name: Optional[Union[str, int]] = 0,
                  has_header: bool = True,
                  use_pandas: bool = True) -> Union[List[Dict[str, Any]], pd.DataFrame]:
        """
        Read an Excel file into a list of dictionaries or a pandas DataFrame.
        
        Args:
            file_path (str): Path to the Excel file
            sheet_name (Union[str, int], optional): Sheet name or index (default: 0)
            has_header (bool): Whether the sheet has a header row
            use_pandas (bool): Whether to use pandas if available (default: True)
            
        Returns:
            Union[List[Dict[str, Any]], pd.DataFrame]: Parsed data
        """
        if use_pandas and HAS_PANDAS:
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=0 if has_header else None)
                return df
            except Exception as e:
                logger.warning(f"Error reading Excel with pandas, falling back to openpyxl: {str(e)}")
        
        # Fallback to openpyxl if available
        if HAS_OPENPYXL:
            try:
                wb = load_workbook(file_path, read_only=True)
                
                # Get the specified sheet
                if isinstance(sheet_name, int):
                    sheet = wb.worksheets[sheet_name]
                else:
                    sheet = wb[sheet_name]
                
                data = []
                rows = list(sheet.rows)
                
                if has_header and rows:
                    # Use first row as header
                    header = [cell.value for cell in rows[0]]
                    
                    # Create dictionaries for each row
                    for row in rows[1:]:
                        row_values = [cell.value for cell in row]
                        row_dict = dict(zip(header, row_values))
                        data.append(row_dict)
                else:
                    # No header, just return values as lists
                    for row in rows:
                        row_values = [cell.value for cell in row]
                        data.append(row_values)
                
                return data
            except Exception as e:
                logger.exception(f"Error reading Excel file {file_path}: {str(e)}")
                raise
        else:
            raise ImportError("Neither pandas nor openpyxl is available for Excel parsing")
    
    @staticmethod
    def read_json(file_path: str) -> Any:
        """
        Read a JSON file into a Python object.
        
        Args:
            file_path (str): Path to the JSON file
            
        Returns:
            Any: Parsed JSON data
        """
        try:
            with open(file_path, 'r', encoding='utf-8') as json_file:
                return json.load(json_file)
        except Exception as e:
            logger.exception(f"Error reading JSON file {file_path}: {str(e)}")
            raise
    
    @staticmethod
    def read_xml(file_path: str, xpath: Optional[str] = None) -> Any:
        """
        Read an XML file into a Python object.
        
        Args:
            file_path (str): Path to the XML file
            xpath (str, optional): XPath query to extract specific elements
            
        Returns:
            Any: Parsed XML data (ElementTree object or elements matching xpath)
        """
        if not HAS_XML:
            raise ImportError("XML parsing is not available")
            
        try:
            tree = ET.parse(file_path)
            root = tree.getroot()
            
            if xpath:
                return root.findall(xpath)
            else:
                return tree
        except Exception as e:
            logger.exception(f"Error reading XML file {file_path}: {str(e)}")
            raise
    
    @staticmethod
    def read_geospatial(file_path: str) -> Any:
        """
        Read a geospatial file into a GeoDataFrame or Python object.
        
        Supports various formats including GeoJSON, Shapefile, GeoPackage, etc.
        
        Args:
            file_path (str): Path to the geospatial file
            
        Returns:
            Any: Parsed geospatial data
        """
        if not HAS_GEOPANDAS:
            raise ImportError("GeoPandas is not available for geospatial parsing")
            
        try:
            gdf = gpd.read_file(file_path)
            return gdf
        except Exception as e:
            logger.exception(f"Error reading geospatial file {file_path}: {str(e)}")
            raise
    
    @staticmethod
    def write_csv(data: Union[List[Dict[str, Any]], pd.DataFrame], 
                 file_path: str,
                 delimiter: str = ',',
                 encoding: str = 'utf-8',
                 index: bool = False) -> str:
        """
        Write data to a CSV file.
        
        Args:
            data (Union[List[Dict[str, Any]], pd.DataFrame]): Data to write
            file_path (str): Path to the output CSV file
            delimiter (str): Field delimiter (default: ',')
            encoding (str): File encoding (default: 'utf-8')
            index (bool): Whether to include index column for pandas DataFrame
            
        Returns:
            str: Path to the written file
        """
        # Check if data is a pandas DataFrame
        if HAS_PANDAS and isinstance(data, pd.DataFrame):
            try:
                data.to_csv(file_path, sep=delimiter, encoding=encoding, index=index)
                return file_path
            except Exception as e:
                logger.warning(f"Error writing DataFrame to CSV, falling back to csv module: {str(e)}")
        
        # Fallback to standard csv module for list of dictionaries
        try:
            if isinstance(data, list) and len(data) > 0:
                with open(file_path, 'w', encoding=encoding, newline='') as csvfile:
                    if isinstance(data[0], dict):
                        fieldnames = data[0].keys()
                        writer = csv.DictWriter(csvfile, fieldnames=fieldnames, delimiter=delimiter)
                        writer.writeheader()
                        writer.writerows(data)
                    else:
                        writer = csv.writer(csvfile, delimiter=delimiter)
                        writer.writerows(data)
                return file_path
            else:
                raise ValueError("Data must be a non-empty list or DataFrame")
        except Exception as e:
            logger.exception(f"Error writing to CSV file {file_path}: {str(e)}")
            raise
    
    @staticmethod
    def write_json(data: Any, file_path: str, indent: int = 2) -> str:
        """
        Write data to a JSON file.
        
        Args:
            data (Any): Data to write
            file_path (str): Path to the output JSON file
            indent (int): Indentation level (default: 2)
            
        Returns:
            str: Path to the written file
        """
        try:
            with open(file_path, 'w', encoding='utf-8') as json_file:
                json.dump(data, json_file, indent=indent, default=str)
            return file_path
        except Exception as e:
            logger.exception(f"Error writing to JSON file {file_path}: {str(e)}")
            raise
    
    @staticmethod
    def write_excel(data: Union[Dict[str, Any], List[Dict[str, Any]], pd.DataFrame], 
                   file_path: str,
                   sheet_name: str = 'Sheet1',
                   index: bool = False) -> str:
        """
        Write data to an Excel file.
        
        Args:
            data (Union[Dict[str, Any], List[Dict[str, Any]], pd.DataFrame]): Data to write
            file_path (str): Path to the output Excel file
            sheet_name (str): Name of the sheet (default: 'Sheet1')
            index (bool): Whether to include index column for pandas DataFrame
            
        Returns:
            str: Path to the written file
        """
        if not HAS_PANDAS and not HAS_OPENPYXL:
            raise ImportError("Neither pandas nor openpyxl is available for Excel writing")
            
        try:
            if HAS_PANDAS:
                # Convert to DataFrame if needed
                if isinstance(data, list):
                    df = pd.DataFrame(data)
                elif isinstance(data, dict):
                    df = pd.DataFrame([data])
                else:
                    df = data
                
                # Create a writer and write the DataFrame
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=index)
                
                return file_path
            elif HAS_OPENPYXL:
                # Fallback to openpyxl if pandas is not available
                from openpyxl import Workbook
                
                wb = Workbook()
                ws = wb.active
                ws.title = sheet_name
                
                # Convert data to list of dictionaries if needed
                if isinstance(data, dict):
                    rows = [data]
                elif isinstance(data, list):
                    rows = data
                else:
                    raise ValueError("Data must be a dictionary, list of dictionaries, or DataFrame")
                
                # Write header row
                if rows:
                    header = list(rows[0].keys())
                    ws.append(header)
                    
                    # Write data rows
                    for row in rows:
                        ws.append([row.get(key) for key in header])
                
                wb.save(file_path)
                return file_path
        except Exception as e:
            logger.exception(f"Error writing to Excel file {file_path}: {str(e)}")
            raise
    
    @staticmethod
    def convert_file(input_path: str, output_path: str, file_format: str) -> str:
        """
        Convert a file from one format to another.
        
        Args:
            input_path (str): Path to the input file
            output_path (str): Path to the output file
            file_format (str): Target format ('csv', 'json', 'excel', 'xml', 'geojson')
            
        Returns:
            str: Path to the converted file
        """
        # Determine input format from file extension
        input_ext = os.path.splitext(input_path)[1].lower()
        
        # Read the input file
        if input_ext in ['.csv', '.txt']:
            data = FileParser.read_csv(input_path)
        elif input_ext in ['.xlsx', '.xls']:
            data = FileParser.read_excel(input_path)
        elif input_ext in ['.json', '.geojson']:
            data = FileParser.read_json(input_path)
        elif input_ext in ['.xml']:
            data = FileParser.read_xml(input_path)
        elif input_ext in ['.shp', '.gpkg', '.kml']:
            data = FileParser.read_geospatial(input_path)
        else:
            raise ValueError(f"Unsupported input format: {input_ext}")
        
        # Write to the output format
        if file_format.lower() == 'csv':
            return FileParser.write_csv(data, output_path)
        elif file_format.lower() == 'json':
            return FileParser.write_json(data, output_path)
        elif file_format.lower() == 'excel':
            return FileParser.write_excel(data, output_path)
        elif file_format.lower() == 'geojson' and HAS_GEOPANDAS:
            if not isinstance(data, gpd.GeoDataFrame):
                raise ValueError("Data must be a GeoDataFrame for GeoJSON output")
            data.to_file(output_path, driver='GeoJSON')
            return output_path
        else:
            raise ValueError(f"Unsupported output format: {file_format}")